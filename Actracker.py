##################
### BIBLIOTECAS ##
##################

import pandas as pd
import datetime as dt 
import time  
import tkinter as tk 
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side
import pygame
import random
import sys
from pathlib import Path
import os



##################
### PATHS ########
##################

# Path del código.
CODE_PATH = Path(__file__).parent if '__file__' in globals() else Path(os.getcwd())

# Path del excel.
EXCEL_PATH = CODE_PATH / 'Periods.xlsx'

# Path to your sound file.
SOUND_PATH = CODE_PATH / 'Alarm.mp3'



##################
### FRASES #######
##################

PHRASES = [
            "Seguí el plan. Confiá en el yo de ayer que te encomendó lo de hoy.",
            "Pequeñitos esfuerzos repetidos día tras día hacen al gran resultado.",
            "No es que tengamos poco tiempo, sino que perdemos mucho. No lo pierdas.",
            "La adversidad es una oportunidad para la virtud.",
            "Hoy es un buen día para hacer las cosas bien.",
            "Si racionalizás ahora y dejás el deber para después, ¿qué te impide en un rato, y mañana, y siempre?",
            "La próxima decisión que tomes es el reflejo de lo que sos y vas a ser. Está en tus manos.",
            "El discomfort es el fuego que forja el carácter.",
            "Esfuerzos continuos y chiquitos son más poderosos que ráfagas intensas pero despedigadas.",
            "Recuerdá: lo que te pide el cuerpo no siempre es lo que necesitás.",
            "Intensificá ahora el hábito de posponer la gratificación espontánea",
            "Si cumplís tu encomienda actual, estás hecho. Lo demás está planeado.",
            "Preocupate solo por la hora que tenés por delante.",
            "Disciplina igual a libertad.",
            "Solo tres meses más.",
            "Hacelo ahora. A veces 'después' se convierte en 'nunca'."
            "Dejá de hablar de lo que es ser un hombre y sé uno.",
            "Es una de dos: el dolor de la disciplina o el dolor del lamento.",
            "Poseer autodisciplina y fuerza de voluntad es tener la capacidad de hacer cosas difíciles o desagradables porque dichas acciones favorecen tu bienestar a largo plazo.",
            "Una vida carente de disciplina está necesariamente llena de remordimientos.",
            "Ejercitá tu autocontrol.",
            "Una voz te dice que no es posible, que no podés: ponela a prueba.",
            "Sumergite en la tarea. Olvidate de lo externo, de tu cabeza, de lo que digas.",
            "Fortificar el hábito de no posponer lo que hay que hacer, de hacerlo ya.",
            f"La regla del 40% dice que cuando la mente de un individuo empieza a decirle que está física o emocionalmente agotado, en realidad solo ha empleado un 40% de su capacidad.",
            "Escuchás una voz en tu cabeza que te dice que estás muy cansado o débil para seguir. Si le das una pausa a esa voz y te esforzás por hacer un poco más, le restás crédito a su tesis. Y, además, cae en consideración, porque ya no le vas a dar tanta bola, como a alguien que habla al pedo.",
            "Tenete fe en que podés hacerlo. Otros días mucho más cansado de lo que estás hoy has podido mucho más.",
            "Pensá en la relajación del momento en que termines, sin reproches ni pendientes.",
            "El dolor es parte del proceso. Soportalo como un hombre.",
            "La incomodidad es el grito de tu parte floja, del hombrecito que te quiere todo para él para consumirte. No lo escuches y seguí.",
            "Ponete a prueba y estudiá hasta qué punto sos capaz de hacer en un día.",
            "Esto es un experimento científico que tiene el fin de estudiar hasta qué punto podés resistir.",
            "Vos elegiste hacer esto. Recordá el porqué. No es tedio, tiene sentido. Introducite en el sentido, sé él.",
            "Cuestioná la legitimidad de las excusas que te estás dando. Anotalas y sometelas a juicio.",
            "La regla de 10X establece que debes establecer objetivos 10 veces superiores a lo que pensás que querés y luego invertir 10 veces más acción de la que pensás que es necesaria para lograr esos objetivos.",
            "Ejecutá el plan con disciplina y esfuerzo.",
            "Si sentís el deseo de abandonar, continuá 10 minutos. Es poquito.",
            "Recordá el tipo del cincel y la piedra. Sé ese tipo hoy. Recordá tus porqués.",
            "No esperés más para “estar listo”, o “sentirte preparado”.",
            "La inacción va de la mano con las excusas.",
            "Cuando todo se siente cómodo y preparado ya es demasiado tarde, habrás esperado demasiado tiempo.",
            "Cada vez que esperás a que las circunstancias mejoren te estás diciendo que sos incapaz en el momento presente.",
            "Podés planificar para el futuro y demorar el inicio todo lo que quieras, pero lo mejor que podés hacer es empezar.",
            "La disciplina solo surge mediante acciones consistentes.",
            "Sé consistente. Empezá ahora.",
            "Pensá en el escenario de vos haciendo todo lo que te falta. Es posible, de lo único que depende es de que sientes el culo y te pongas a laburar más que el resto.",
            "Sin importar lo mucho que pienses que merecés un descanso o una recompensa por tus acciones pasadas, aceptar esa tendencia a racionalizar o dar excusas va a tener un efecto negativo sobre tu disciplina.",
            "La recompensa por una buena acción es haberla hecho.",
            "Lo que necesitás en abundancia no es la autodisciplina por sí misma, sino una gran capacidad para manejar y tolerar la incomodidad.",
            "De la misma manera que levantar pesas produce molestias pasajeras que contribuyen a fortalecer tu musculatura, actuar de manera disciplinada y tomar las decisiones correctas también fortalece tu “músculo de incomodidad”.",
            "Convertir en un hábito cotidiano el abrazar situaciones incómodas puede tener un efecto positivo en todos los aspectos de tu vida.",
            "A la disciplina no le importa que estés exhausto, irritado o abatido: esos son los momentos cuando más la necesitás.",
            "Compará los impulsos con olas que puedes surfear: ascienden en intensidad, llegan a un pico y, eventualmente, rompen.",
            "Sentís un impulso. Detenete por un momento. Pensá acerca de él. Prestá atención a las emociones y sensaciones físicas que te provoca. Aceptalo. Notá cómo las sensaciones evolucionan con el transcurso del tiempo. Mientras hacés esto, enfocate en tu respiración para ayudarte a sobrellevar el deseo, imaginá que es una ola y surfeás a través de ella.",
            "Luchar contra los impulsos es rara vez efectivo, pero al observarlos con curiosidad, sin identificarte con ellos, tenés más probabilidades de superarlos.",
            "El poder de tus impulsos proviene de tu disposición a consentirlos.",
            "Los impulsos como una cascada, y batallarlos sería equivalente a tratar de bloquear la caída de agua. Por supuesto, es inevitable que la cascada rompa la barrera, quizás incluso con más fuerza, debido a la presión contenida mientras la bloqueabas. El mindfulness es el escape a esta situación imposible, porque en lugar de tratar de bloquear la cascada, o el deseo, das un paso atrás y solo lo observás.",
            "Estudiá tus hábitos de manera científica.",
            "Seguí esforzándote hasta que esto sea tan natural como respirar.",
            "Las grandes gestas son difíciles, no pueden ser fáciles. Hacelo y contátelo al final del día con orgullo.",
            "Date herramientas para sentirte orgulloso de vos mismo. ¿O cuáles son las que te enorgullecen ahora? Las que costaron trabajo y dedicación."
]   



##################
### VARIABLES ####
##################

# Columnas del df.
INICIO = 'Inicio'
FINAL = 'Final'
PLAN_PREVISTO = 'Plan_Previsto'
ACTIVIDAD_REALIZADA = 'Actividad_Realizada'
EXPLICACION = 'Explicación'

# Variables.
MINUTES_PERIOD = 10
TIME_ERROR = 1

# Noche.
START_NIGHT = dt.time(22,0)
END_NIGHT = dt.time(6,0)

# Hora actual.
START_TIME = dt.datetime.now()



##################
### FUNCIONES ####
##################

def Play_Sound_For_Duration(File_Path: str, Duration_Seconds: float) -> None:

    """
    Plays a sound from a given file for a specified duration.
    
    Args:
    File_Path (str): The path to the sound file.
    Duration_Seconds (float): Duration in seconds for which the sound should play.
    
    Returns:
    None

    """

    # Initialize pygame mixer.
    pygame.mixer.init()

    # Load the sound file.
    pygame.mixer.music.load(File_Path)

    # Play the sound.
    pygame.mixer.music.play()

    # Wait for the specified duration.
    time.sleep(Duration_Seconds)

    # Stop the sound after the duration has passed.
    pygame.mixer.music.stop()

def Add_Row_To_DataFrame(Row: dict, df: pd.DataFrame, Fill: str | int | bool | float = 0, 
                         Last = True) -> pd.DataFrame:

    """
    Adds a new row to a specified DataFrame based on a dictionary.

    This function creates a new row in a DataFrame using the values 
    provided in a dictionary. If a key in the dictionary does not 
    match a column in the DataFrame, that column will be filled 
    with the specified value. The new row can be added at the end 
    or the beginning of the DataFrame.

    Parameters:
    -----------
    Row : dict
        A dictionary containing the data for the new row. The keys 
        should match the column names in the DataFrame.

    df : pd.DataFrame
        The DataFrame to which the new row will be added.

    Last : bool, optional
        If True (default), the new row is added to the end of the 
        DataFrame. If False, it is added to the beginning.

    Fill : str, int, bool or float, optional
        The value used to fill in any missing columns in the new 
        row if the corresponding keys are not found in the dictionary. 
        The default is 0. This can be any valid data type.

    Returns:
    --------
    pd.DataFrame
        The updated DataFrame with the new row added.
    
    Notes:
    ------
    - The function checks if each key in the dictionary matches a 
      column in the DataFrame. If not, the Fill value is assigned 
      to that column for the new row.
    - The DataFrame index is reset after adding the new row.

    Example:
    ---------
    >>> data = {'Name': 'David'}
    >>> updated_df = Add_Row_To_DataFrame(data, df, Fill='Unknown')

    """
    
    New_Row = pd.DataFrame()

    for Column in df.columns:
        if Column in list(Row.keys()):
            New_Row.loc[0, Column] = Row[Column]
        else:
            New_Row.loc[0, Column] = Fill

    if Last:
        df = pd.concat([df, New_Row], ignore_index=True)
    else:
        df = pd.concat([New_Row, df], ignore_index=True)

    df = df.reset_index(drop=True)

    return df

def Adjust_Column_Width(Path: str, Min_Column: int, Max_Column: int, 
                         Width: float, Sheet_Name = None):
    
    """
    Adjust the width of specified columns in a worksheet.

    Parameters:
    - Path (str): The path to the workbook file.
    - Min_Column (int): The starting column index (1-based).
    - Max_Column (int): The ending column index (1-based).
    - Width (float): The width to set for the columns.
    - Sheet_Name (str, optional): The name of the sheet to modify.
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
        
    for Column in range(Min_Column, Max_Column + 1):
        Sheet.column_dimensions[chr(64 + Column)].width = Width  # type: ignore
    
    Book.save(Path)

def Formating_Book(Path: str, Sheet_Name = None):

    """
    Format a workbook by applying filters, alignment, and borders.

    Parameters:
    - Path (str): The path to the workbook file.
    - Sheet_Name (str, optional): The name of the sheet to format. 
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active  
    else:
        if Sheet_Name not in Book.sheetnames:
            raise ValueError(f"Sheet '{Sheet_Name}' does not exist in the workbook.")
        Sheet = Book[Sheet_Name]

    # Apply autofilter.
    Sheet.auto_filter.ref = Sheet.dimensions # type: ignore

    Align_Format = Alignment(horizontal='center', vertical='center')
    Border_Format = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    for Column in range(1, Sheet.max_column + 1): # type: ignore
        for Row in range(1, Sheet.max_row + 1): # type: ignore
            Cell = Sheet.cell(row=Row, column=Column) # type: ignore
            Cell.alignment = Align_Format
            Cell.border = Border_Format

    Book.save(Path)

def Disable_Event():
    pass  

def Press_Done():
    global df, Last_Row_Index
        
    df.loc[Last_Row_Index, ACTIVIDAD_REALIZADA] = '✓'
    df.loc[Last_Row_Index, EXPLICACION] = '-'
    Open_Promise_Window()

def Press_Undone():
    global Window
    Open_Activity_Window()

def Open_Activity_Window():
    global Window, df, Last_Row_Index
    Activity_Window = tk.Toplevel(Window) 
    Activity_Window.title("Actividad realizada")
    Activity_Window.geometry("1200x500")
    Activity_Window.geometry("+{}+{}".format(int(Activity_Window.winfo_screenwidth() / 2 - 600), int(Activity_Window.winfo_screenheight() / 2 - 250)))  # Centrar la ventana.

    Activity_Window.attributes('-topmost', True)  # Mantener ventana al frente.
    Activity_Window.attributes('-toolwindow', True)  # Deshabilitar minimizar. 

    Activity_Window.transient(Window)  
    Activity_Window.focus_force()  # Forzar el foco en la ventana.
    Activity_Window.grab_set()  # Impedir hacer clic fuera de la ventana.   
    Activity_Window.protocol("WM_DELETE_WINDOW", Disable_Event)

    def Set_Initial_Position():
        global Initial_X, Initial_Y
        Initial_X = Activity_Window.winfo_x()
        Initial_Y = Activity_Window.winfo_y()
    
    # Establecer la posición inicial después de 10ms para asegurar que la ventana esté completamente renderizada.
    Activity_Window.after(10, Set_Initial_Position)

    # Definir función para bloquear el movimiento de la ventana.
    def Block_Movement(event):
        Activity_Window.geometry(f"+{Initial_X}+{Initial_Y}")

    # Vincular el evento de movimiento a la función Block_Movement.
    Activity_Window.bind("<Configure>", Block_Movement)

    Label = tk.Label(Activity_Window, text="¿Qué hiciste en lugar de lo previsto?", font=("Calibri Light", 14))
    Label.pack(pady=20)

    Activity_Box = tk.Entry(Activity_Window, width=30, font=("Calibri Light", 14), justify='center')  # Centrar cursor en la caja de entrada.
    Activity_Box.pack(pady=10, padx=20, expand=True, fill='both')

    Label = tk.Label(Activity_Window, text="¿Por qué?", font=("Calibri Light", 14))
    Label.pack(pady=20)

    Justify_Box = tk.Entry(Activity_Window, width=30, font=("Calibri Light", 14), justify='center')
    Justify_Box.pack(pady=10, padx=20, expand=True, fill='both')

    def Press_Close():
        global df, Last_Row_Index

        if not Activity_Box.get().strip():
            tk.messagebox.showwarning("Empty Field", "Por favor, ingresa una actividad realizada.")
            return
        if not Justify_Box.get().strip():
            tk.messagebox.showwarning("Empty Field", "Por favor, ingresa una explicación.")
            return

        df.loc[Last_Row_Index, ACTIVIDAD_REALIZADA] = Activity_Box.get()
        df.loc[Last_Row_Index, EXPLICACION] = Justify_Box.get()
        Activity_Window.destroy()
        Open_Promise_Window()

    Close_Button = tk.Button(Activity_Window, text="Close", command=Press_Close, font=("Calibri Light", 14))
    Close_Button.pack(pady=20)

    Window.wait_window(Activity_Window)

def Open_Promise_Window():
    global Window, MINUTES_PERIOD, Phrase, Promise_Text  # Declarar Promise_Text como global.

    Promise_Window = tk.Toplevel(Window) 
    Promise_Window.title("Planear próximo intervalo")
    Promise_Window.geometry("1200x500")
    Promise_Window.geometry("+{}+{}".format(int(Promise_Window.winfo_screenwidth() / 2 - 600), int(Promise_Window.winfo_screenheight() / 2 - 250)))  # Centrar ventana.

    Promise_Window.attributes('-topmost', True)  # Mantener la ventana al frente.
    Promise_Window.attributes('-toolwindow', True)  # Deshabilitar minimizar.
    Promise_Window.transient(Window)  
    Promise_Window.focus_force()  # Forzar el foco en la ventana.
    Promise_Window.grab_set()  # Impedir clic fuera de la ventana.
    Promise_Window.protocol("WM_DELETE_WINDOW", Disable_Event)

    # Definir función para bloquear el movimiento de la ventana.
    def Block_Movement(event):
        Promise_Window.geometry(f"+{Initial_X}+{Initial_Y}")

    # Vincular el evento de movimiento a la función Block_Movement.
    Promise_Window.bind("<Configure>", Block_Movement)

    # Elementos de la ventana.
    Label = tk.Label(Promise_Window, text=f"¿Qué vas a hacer en los próximos {MINUTES_PERIOD} minutos?", font=("Calibri Light", 14))
    Label.pack(pady=20)

    Promise_Box = tk.Entry(Promise_Window, width=30, font=("Calibri Light", 14), justify='center')
    Promise_Box.pack(pady=10, padx=20, expand=True, fill='both')

    def Press_Promise():
        global df, Last_Row_Index, MINUTES_PERIOD, End_Last_Period, Now_Minus_Period, Difference_Minutes
        global INICIO, FINAL, PLAN_PREVISTO, ACTIVIDAD_REALIZADA, EXPLICACION, Promise_Text

        Promise_Text = Promise_Box.get().strip()
        
        if not Promise_Text: 
            tk.messagebox.showwarning("Empty Field", "Por favor, ingresa una actividad para continuar.")
            return

        Period = {
            INICIO: START_TIME,
            FINAL: START_TIME + dt.timedelta(days=0, hours=0, minutes=MINUTES_PERIOD),
            PLAN_PREVISTO: Promise_Text,
            ACTIVIDAD_REALIZADA: '-',
            EXPLICACION: '-'
        }

        df = Add_Row_To_DataFrame(Period, df, Fill='-')
        Window.destroy()

    # Frase y botón de "Prometo".
    Prhase_Label = tk.Label(Promise_Window, text = f'"{Phrase}"', font=("Calibri Light", 14, "italic"), wraplength=800)
    Prhase_Label.pack(pady=20)

    Promise_Button = tk.Button(Promise_Window, text = "Prometo", command = Press_Promise, font = ("Calibri Light", 14))
    Promise_Button.pack(pady=20)

    Window.wait_window(Promise_Window)

def Check_Time(Window, Start, Seconds):
    global Promise_Text

    if (dt.datetime.now() - Start).total_seconds() > Seconds:
        global df, Last_Row_Index, MINUTES_PERIOD, End_Last_Period, Now_Minus_Period, Difference_Minutes
        global INICIO, FINAL, PLAN_PREVISTO, ACTIVIDAD_REALIZADA, EXPLICACION

        # Period = {
        #     INICIO: START_TIME,
        #     FINAL: START_TIME + dt.timedelta(days=0, hours=0, minutes=MINUTES_PERIOD),
        #     PLAN_PREVISTO: '-',
        #     ACTIVIDAD_REALIZADA: '-',
        #     EXPLICACION: '-'
        # }

        # df = Add_Row_To_DataFrame(Period, df, Fill='-')

        # # Reemplazar cualquier valor no válido (como '-') por NaT antes de convertir a datetime.
        # df[FINAL] = pd.to_datetime(df[FINAL].replace('-', pd.NaT), format='%Y-%m-%d %H:%M', errors='coerce') # type: ignore

        # # Guardar en Excel.
        # df[INICIO] = pd.to_datetime(df[INICIO]).dt.strftime('%Y-%m-%d %H:%M')  
        # df[FINAL] = pd.to_datetime(df[FINAL]).dt.strftime('%Y-%m-%d %H:%M')    
        # df.to_excel(EXCEL_PATH.as_posix(), index=False)

        # Adjust_Column_Width(EXCEL_PATH.as_posix(), 1, 2, 25)
        # Adjust_Column_Width(EXCEL_PATH.as_posix(), 3, 5, 40)
        # Formating_Book(EXCEL_PATH.as_posix())

        Window.destroy()
        sys.exit()
    else:
        Window.after(2000, Check_Time, Window, Start, Seconds)

##################
### PROGRAMA #####
##################

# Call the function to play sound for 2 seconds.
Play_Sound_For_Duration(SOUND_PATH.as_posix(), 0)

# Cargar base de datos
df = pd.read_excel(EXCEL_PATH.as_posix())

# Índice de la última fila
if len(df) == 1:
    Last_Row_Index = 0
else:
    Last_Row_Index = len(df) - 1

# Calcular el tiempo desde ahora hacia atrás con el período elegido.
Now_Minus_Period = START_TIME - dt.timedelta(minutes=MINUTES_PERIOD)

# Si el df no está vacío, busca el final del período en la última fila.
if len(df) > 0:
    End_Last_Period = df.loc[Last_Row_Index, FINAL]  # Obtener el final del período
    
    # Verificar si End_Last_Period es un objeto Timestamp
    if isinstance(End_Last_Period, pd.Timestamp):
        End_Last_Period = End_Last_Period.to_pydatetime() 
    elif isinstance(End_Last_Period, str):
        try:
            End_Last_Period = dt.datetime.strptime(End_Last_Period, '%Y-%m-%d %H:%M')
        except ValueError as e:
            print(f"Error al convertir la cadena a datetime: {e}")  
            End_Last_Period = None
            
    else:
        print(f"Tipo de dato inesperado: {type(End_Last_Period)}")    
else:
    End_Last_Period = Now_Minus_Period

# Definir diferencia de horario.    
Difference = Now_Minus_Period - End_Last_Period # type: ignore
Difference_Minutes = round(Difference.total_seconds() / 60) # type: ignore

# Check if the time is between 22:00 and 06:00.
if START_TIME.time() >= START_NIGHT or START_TIME.time() <= END_NIGHT:
    Phrase = 'Es hora de irse a dormir.'
else:
    Phrase = random.choice(PHRASES)

if Difference_Minutes <= TIME_ERROR and len(df) > 0:

    Window = tk.Tk()
    Window.title(f"Pasaron los {MINUTES_PERIOD} minutos...")
    Window.geometry("1200x500")
    Window.geometry("+{}+{}".format(int(Window.winfo_screenwidth() / 2 - 600), int(Window.winfo_screenheight() / 2 - 250)))  # Center the window.

    # Atributos para mantener la ventana centrada y bloqueada.
    def Set_Initial_Position():
        global Initial_X, Initial_Y
        Initial_X = Window.winfo_x()
        Initial_Y = Window.winfo_y()

    Window.after(10, Set_Initial_Position)  # Establecer posición inicial después de 10 ms.

    def Block_Movement(event):
        Window.geometry(f"+{Initial_X}+{Initial_Y}")

    Window.bind("<Configure>", Block_Movement)

    Plan_Label = tk.Label(Window, text="Este era tu plan:", font=("Calibri Light", 14))
    Plan_Label.pack(pady=20)

    Bold_Label = tk.Label(Window, text=df[PLAN_PREVISTO][Last_Row_Index], font=("Calibri Light", 14, "bold"))
    Bold_Label.pack(pady=20)

    Question_Label = tk.Label(Window, text="¿Lo hiciste?", font=("Calibri Light", 14))
    Question_Label.pack(pady=20)

    Button_Frame = tk.Frame(Window)
    Button_Frame.pack(pady=20)

    Yes_Button = tk.Button(Button_Frame, text="Sí", command=Press_Done, font=("Calibri Light", 14), width=10, height=1)
    Yes_Button.pack(side=tk.LEFT, padx=5)

    No_Button = tk.Button(Button_Frame, text="No", command=Press_Undone, font=("Calibri Light", 14), width=10, height=1)
    No_Button.pack(side=tk.LEFT, padx=5)

    Window.attributes('-topmost', True)
    Window.attributes('-toolwindow', True)
    Window.focus_force()
    Window.grab_set()
    Window.protocol("WM_DELETE_WINDOW", Disable_Event)

    Check_Time(Window, START_TIME, 540)
    Window.mainloop()

else:
    Window = tk.Tk()
    Window.title("Se inicia un nuevo período.")
    Window.geometry("1200x500")
    Window.geometry("+{}+{}".format(int(Window.winfo_screenwidth() / 2 - 600), int(Window.winfo_screenheight() / 2 - 250)))  # Center the window.

    # Atributos para mantener la ventana centrada y bloqueada.
    def Set_Initial_Position():
        global Initial_X, Initial_Y
        Initial_X = Window.winfo_x()
        Initial_Y = Window.winfo_y()

    Window.after(10, Set_Initial_Position)  # Establecer posición inicial después de 10 ms.

    def Block_Movement(event):
        Window.geometry(f"+{Initial_X}+{Initial_Y}")

    Window.bind("<Configure>", Block_Movement)

    Start_Label = tk.Label(Window, text="Arranque, maestro.", font=("Calibri Light", 14))
    Start_Label.pack(pady=20)

    def Press_Start():
        Open_Promise_Window()

    Start_Button = tk.Button(Window, text="Empezar", command=Press_Start, font=("Calibri Light", 14))
    Start_Button.pack(pady=10)

    Window.attributes('-topmost', True)
    Window.attributes('-toolwindow', True)
    Window.focus_force()
    Window.grab_set()
    Window.protocol("WM_DELETE_WINDOW", Disable_Event)

    for i in range(120):
        if (dt.datetime.now() - START_TIME).total_seconds() > 120:
            exit()

    Check_Time(Window, START_TIME, 540)
    Window.mainloop()
    
# Reemplazar cualquier valor no válido (como '-') por NaT antes de convertir a datetime
df[FINAL] = pd.to_datetime(df[FINAL].replace('-', pd.NaT), format='%Y-%m-%d %H:%M', errors='coerce') # type: ignore

# Guardar en excel.
df[INICIO] = pd.to_datetime(df[INICIO]).dt.strftime('%Y-%m-%d %H:%M')  
df[FINAL] = pd.to_datetime(df[FINAL]).dt.strftime('%Y-%m-%d %H:%M')    
df.to_excel(EXCEL_PATH.as_posix(), index=False)

Adjust_Column_Width(EXCEL_PATH.as_posix(), 1, 2, 25)
Adjust_Column_Width(EXCEL_PATH.as_posix(), 3, 5, 40)
Formating_Book(EXCEL_PATH.as_posix())